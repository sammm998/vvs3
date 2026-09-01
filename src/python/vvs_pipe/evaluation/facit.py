"""Compare a finished blind analysis against a facit.

The comparison never touches the pipeline: it takes an ``AnalysisResult`` that
has already been produced and a ground-truth file, and reports the difference.
Where the engine and the facit disagree the engine's number is reported as it
is - the specification is explicit that a 57.5 m measurement against a 57.2 m
facit is reported as 57.5 m with a +0.3 m difference, never adjusted.

Two facit formats are accepted:

* the JSON manifest emitted by the fixture generator;
* a spreadsheet, read with openpyxl, whose header row names the columns
  (designation / diameter / horizontal / vertical / total in any language, by
  position or by a header keyword).
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence

from ..canonical import ql, qs

LENGTH_TOLERANCE_M = 0.05
LENGTH_RELATIVE_TOLERANCE = 0.01
DIAMETER_TOLERANCE_MM = 2.0

_HEADER_HINTS = {
    "designation": ("designation", "beteckning", "kod", "code", "benamning"),
    "diameter": ("diameter", "dimension", "dn", "dim"),
    "horizontal": ("horizontal", "horisontell", "liggande", "h"),
    "vertical": ("vertical", "vertikal", "stam", "stigare", "v"),
    "total": ("total", "summa", "totalt", "sum"),
}


@dataclass(frozen=True, slots=True)
class TruthRow:
    designation: str
    diameter_mm: float | None
    horizontal_m: float | None
    vertical_m: float | None
    total_m: float | None


def load_ground_truth(path: str | Path) -> tuple[TruthRow, ...]:
    p = Path(path)
    if p.suffix.lower() == ".json":
        return _load_json(p)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        return _load_xlsx(p)
    raise ValueError(f"unsupported ground-truth format: {p.suffix}")


def _load_json(path: Path) -> tuple[TruthRow, ...]:
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = data.get("quantities", data if isinstance(data, list) else [])
    out = [
        TruthRow(
            designation=str(r["designation"]),
            diameter_mm=_num(r.get("diameterMm")),
            horizontal_m=_num(r.get("horizontalM")),
            vertical_m=_num(r.get("verticalM")),
            total_m=_num(r.get("totalM")),
        )
        for r in rows
    ]
    return tuple(sorted(out, key=lambda r: (r.designation, r.diameter_mm or -1.0)))


def _load_xlsx(path: Path) -> tuple[TruthRow, ...]:
    from openpyxl import load_workbook  # imported here, never at package import

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ()
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    idx: dict[str, int] = {}
    for field, hints in _HEADER_HINTS.items():
        for i, h in enumerate(header):
            if any(hint in h for hint in hints):
                idx.setdefault(field, i)
                break
    out: list[TruthRow] = []
    for raw in rows[1:]:
        if not raw or idx.get("designation") is None:
            continue
        designation = raw[idx["designation"]]
        if designation is None or not str(designation).strip():
            continue
        out.append(
            TruthRow(
                designation=str(designation).strip(),
                diameter_mm=_num(_cell(raw, idx.get("diameter"))),
                horizontal_m=_num(_cell(raw, idx.get("horizontal"))),
                vertical_m=_num(_cell(raw, idx.get("vertical"))),
                total_m=_num(_cell(raw, idx.get("total"))),
            )
        )
    return tuple(sorted(out, key=lambda r: (r.designation, r.diameter_mm or -1.0)))


def _cell(row: Sequence[Any], i: int | None) -> Any:
    return None if i is None or i >= len(row) else row[i]


def _num(v: Any) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _key(designation: str, diameter: float | None) -> tuple[str, float]:
    return (designation, round(diameter, 0) if diameter is not None else -1.0)


def compare_with_ground_truth(result, ground_truth_path: str | Path) -> dict[str, Any]:
    truth = load_ground_truth(ground_truth_path)

    # Fold the engine's rows to one entry per (designation, size); the engine
    # deliberately splits a size across rows when part of it is unverified, and
    # the facit has no such split.
    detected: dict[tuple[str, float], dict[str, Any]] = {}
    for q in result.quantities:
        if q.designation is None:
            continue
        k = _key(q.designation, q.diameter_mm)
        d = detected.setdefault(
            k,
            {
                "designation": q.designation,
                "diameterMm": q.diameter_mm,
                "horizontalM": 0.0,
                "verticalM": 0.0,
                "totalM": 0.0,
                "pipeCount": 0,
                "states": set(),
                "reasons": set(),
            },
        )
        d["horizontalM"] += q.horizontal_m or 0.0
        d["verticalM"] += q.vertical_m or 0.0
        d["totalM"] += q.total_m or 0.0
        d["pipeCount"] += q.pipe_count
        d["states"].add(q.state.value)
        d["reasons"].update(r.value for r in q.reasons)

    truth_map = {_key(t.designation, t.diameter_mm): t for t in truth}

    tp: list[dict[str, Any]] = []
    fp: list[dict[str, Any]] = []
    fn: list[dict[str, Any]] = []
    length_rows: list[dict[str, Any]] = []

    for k in sorted(set(detected) | set(truth_map)):
        got = detected.get(k)
        want = truth_map.get(k)
        if got is not None and want is not None:
            tp.append({"designation": want.designation, "diameterMm": want.diameter_mm})
            length_rows.append(_length_row(got, want))
        elif got is not None:
            fp.append({"designation": got["designation"], "diameterMm": got["diameterMm"]})
        elif want is not None:
            fn.append({"designation": want.designation, "diameterMm": want.diameter_mm})

    precision = len(tp) / max(1, len(tp) + len(fp))
    recall = len(tp) / max(1, len(tp) + len(fn))
    f1 = 0.0 if precision + recall == 0 else 2 * precision * recall / (precision + recall)

    detected_names = {d["designation"] for d in detected.values()}
    truth_names = {t.designation for t in truth}
    designation_coverage = len(detected_names & truth_names) / max(1, len(truth_names))

    all_pipes = [pp for p in result.pages for pp in p.physical_pipes]
    unknown = sum(1 for pp in all_pipes if pp.designation is None)
    ambiguous = sum(1 for pp in all_pipes if pp.identity_state.value == "AMBIGUOUS")
    duplicates = result.reconciliation.duplicate_centerlines + result.reconciliation.runs_in_multiple_pipes

    verticals = [v for p in result.pages for v in p.verticals]
    vertical_rows = [r for r in length_rows if (r["truthVerticalM"] or 0.0) > 0.0]
    vertical_accuracy = _accuracy([r["verticalWithinTolerance"] for r in vertical_rows])

    return {
        "schema": "vvs-pipe/post-test/1",
        "groundTruthFile": str(Path(ground_truth_path).name),
        "note": (
            "Post-hoc only. The analysis in this report was produced before this "
            "file was opened; no detected value has been adjusted to match it."
        ),
        "summary": {
            "truePositives": len(tp),
            "falsePositives": len(fp),
            "falseNegatives": len(fn),
            "precision": qs(precision),
            "recall": qs(recall),
            "f1": qs(f1),
            "designationCoverage": qs(designation_coverage),
            "pipeCoverage": qs(len(tp) / max(1, len(truth_map))),
            "measurementAccuracy": qs(_accuracy([r["totalWithinTolerance"] for r in length_rows])),
            "verticalAccuracy": qs(vertical_accuracy),
            "duplicateRate": qs(duplicates / max(1, len(all_pipes))),
            "unknownRate": qs(unknown / max(1, len(all_pipes))),
            "ambiguousRate": qs(ambiguous / max(1, len(all_pipes))),
            "unresolvedVerticals": sum(1 for v in verticals if v.length_m is None),
        },
        "truePositives": tp,
        "falsePositives": fp,
        "falseNegatives": fn,
        "lengths": length_rows,
    }


def _length_row(got: dict[str, Any], want: TruthRow) -> dict[str, Any]:
    def diff(a: float | None, b: float | None) -> float | None:
        if a is None or b is None:
            return None
        return ql(a - b)

    def within(a: float | None, b: float | None) -> bool | None:
        if a is None or b is None:
            return None
        return abs(a - b) <= max(LENGTH_TOLERANCE_M, LENGTH_RELATIVE_TOLERANCE * abs(b))

    return {
        "designation": want.designation,
        "diameterMm": want.diameter_mm,
        "detectedHorizontalM": ql(got["horizontalM"]),
        "truthHorizontalM": want.horizontal_m,
        "horizontalDifferenceM": diff(got["horizontalM"], want.horizontal_m),
        "horizontalWithinTolerance": within(got["horizontalM"], want.horizontal_m),
        "detectedVerticalM": ql(got["verticalM"]),
        "truthVerticalM": want.vertical_m,
        "verticalDifferenceM": diff(got["verticalM"], want.vertical_m),
        "verticalWithinTolerance": within(got["verticalM"], want.vertical_m),
        "detectedTotalM": ql(got["totalM"]),
        "truthTotalM": want.total_m,
        "totalDifferenceM": diff(got["totalM"], want.total_m),
        "totalWithinTolerance": within(got["totalM"], want.total_m),
        "detectedDiameterMm": got["diameterMm"],
        "diameterWithinTolerance": (
            None
            if got["diameterMm"] is None or want.diameter_mm is None
            else abs(got["diameterMm"] - want.diameter_mm) <= DIAMETER_TOLERANCE_MM
        ),
        "states": sorted(got["states"]),
        "reasons": sorted(got["reasons"]),
    }


def _accuracy(flags: Iterable[bool | None]) -> float:
    vals = [f for f in flags if f is not None]
    return 0.0 if not vals else sum(1 for v in vals if v) / len(vals)
